from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, status
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
from dotenv import load_dotenv
from pymongo import MongoClient
from bson import ObjectId
from bson.errors import InvalidId
from pathlib import Path
import os
import re
import uuid
import pandas as pd
from openpyxl import load_workbook
 
# Import custom modules
from utils import data_ingestion, test_case_utils, user_story_utils
from utils.llms import Mistral, openai, llama
 
# ----------------- FastAPI App Setup -----------------
app = FastAPI()
 
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
 
# ----------------- MongoDB Setup -----------------
load_dotenv()
 
mongo_client = MongoClient(os.getenv("MONGODB_URL", "mongodb://localhost:27017/"))
db = mongo_client["Gen_AI"]
collection = db["test_case_generation"]
 
# ----------------- Directories Setup -----------------
PROMPT_FILE_PATH = os.getenv("PROMPT_FILE_PATH")
USER_STORY_PROMPT_FILE_PATH = os.getenv("USER_STORY_PROMPT_FILE_PATH")
INPUT_DIR = os.getenv("INPUT_DIR", "input_pdfs")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "output_files")
EXCEL_OUTPUT_DIR = os.getenv("EXCEL_OUTPUT_DIR", "excel_files")
 
Path(INPUT_DIR).mkdir(parents=True, exist_ok=True)
Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
Path(EXCEL_OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
 
DEFAULT_CHUNK_SIZE = 7000
TEST_CASES_CACHE = {}
 
# ----------------- Model Dispatcher -----------------
MODEL_DISPATCHER = {
    "Mistral": Mistral.generate_with_mistral,
    "Openai": openai.generate_with_openai,
    "Llama": llama.generate_with_llama,
}
 
# ----------------- Helper Functions -----------------
def split_text_into_chunks(text, chunk_size=7000):
    chunks = []
    current_chunk = ""
    sentences = re.split(r"(?<!\w\.\w.)(?<![A-Z][a-z]\.)\s", text)
    for sentence in sentences:
        if len(current_chunk) + len(sentence) + 1 <= chunk_size:
            current_chunk += sentence + " "
        else:
            chunks.append(current_chunk.strip())
            current_chunk = sentence + " "
    if current_chunk:
        chunks.append(current_chunk.strip())
    return chunks
 
def serialize_document(doc):
    doc["_id"] = str(doc["_id"])
    return doc
 
def clean_stars_from_text(text):
    return text.replace("*", "").strip()
 
def save_text_to_excel(data, output_path, column_names):
    df = pd.DataFrame(data, columns=column_names)
    df.to_excel(output_path, index=False)
 
def add_spacing_in_excel(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active
    rows_to_insert = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower().startswith("expected result"):
                rows_to_insert.append(idx)
    offset = 0
    for row_idx in rows_to_insert:
        ws.insert_rows(row_idx + 1 + offset, amount=2)
        offset += 2
    wb.save(file_path)
 
async def safe_generate(generate_func, *args, **kwargs):
    try:
        return generate_func(*args, **kwargs)
    except Exception as e:
        if "Credit limit exceeded" in str(e):
            raise HTTPException(status_code=402, detail={"message": "Credit limit exceeded.", "error_detail": str(e)})
        else:
            raise HTTPException(status_code=500, detail=str(e))
 
# ----------------- API Endpoints -----------------
 
@app.get("/")
async def root():
    return {"message": "FastAPI backend working! 🚀"}
 
@app.post("/upload-document/")
async def upload_document(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    file_path = Path(INPUT_DIR) / f"{file_id}.pdf"
 
    with open(file_path, "wb") as f:
        f.write(await file.read())
 
    return {"file_id": file_id, "file_path": str(file_path)}
 
@app.post("/process_and_generate/")
async def process_and_generate(
    file: UploadFile = File(...),
    model_name: str = Form("Mistral"),
    chunk_size: Optional[int] = Query(default=None),
    cache_key: Optional[str] = Query(default=None),
):
    if model_name not in MODEL_DISPATCHER:
        raise HTTPException(status_code=400, detail=f"Unsupported model: {model_name}")
 
    generation_function = MODEL_DISPATCHER[model_name]
    file_name = file.filename
    file_path = Path(INPUT_DIR) / file_name
 
    try:
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
    finally:
        await file.close()
 
    brd_text, _ = data_ingestion.load_pdf_text(str(file_path))
    if not brd_text:
        raise HTTPException(status_code=500, detail="Failed to extract text from PDF.")
 
    cleaned_text = data_ingestion.clean_text(brd_text)
    chunk = chunk_size if chunk_size else DEFAULT_CHUNK_SIZE
 
    chunks = split_text_into_chunks(cleaned_text, chunk) if model_name != "Mistral" else [cleaned_text]
 
    all_test_cases, all_user_stories = [], []
    for chunk_text in chunks:
        test_case_text = await safe_generate(
            test_case_utils.generate_test_cases,
            chunk_text, generation_function, PROMPT_FILE_PATH
        )
        if test_case_text:
            all_test_cases.append(clean_stars_from_text(test_case_text))
 
    for chunk_text in chunks:
        user_story_text = await safe_generate(
            user_story_utils.generate_user_stories,
            chunk_text, generation_function, USER_STORY_PROMPT_FILE_PATH
        )
        if user_story_text:
            all_user_stories.append(clean_stars_from_text(user_story_text))
 
    combined_test_cases = "\n".join(all_test_cases)
    combined_user_stories = "\n".join(all_user_stories)
 
    base_stem = Path(file_name).stem
    excel_test_case_path = Path(EXCEL_OUTPUT_DIR) / f"{base_stem}_test_cases.xlsx"
    excel_user_story_path = Path(EXCEL_OUTPUT_DIR) / f"{base_stem}_user_stories.xlsx"
 
    save_text_to_excel([{ "Test Cases": tc } for tc in combined_test_cases.split('\n') if tc.strip()], excel_test_case_path, ["Test Cases"])
    save_text_to_excel([{ "User Stories": us } for us in combined_user_stories.split('\n') if us.strip()], excel_user_story_path, ["User Stories"])
 
    add_spacing_in_excel(excel_test_case_path)
    add_spacing_in_excel(excel_user_story_path)
 
    if not cache_key:
        cache_key = str(uuid.uuid4())
 
    TEST_CASES_CACHE[cache_key] = {
        "test_cases": combined_test_cases,
        "user_stories": combined_user_stories
    }
 
    document = {
        "doc_name": file.filename,
        "doc_path": str(file_path),
        "test_case_excel_path": str(excel_test_case_path),
        "user_story_excel_path": str(excel_user_story_path),
        "selected_model": model_name,
        "llm_response_testcases": combined_test_cases,
        "llm_response_user_stories": combined_user_stories,
    }
    collection.insert_one(document)
 
    return JSONResponse(content={
        "message": "File uploaded and processed successfully",
        "cache_key": cache_key,
        "model_used": model_name
    })
 
@app.get("/documents/")
def get_all_documents():
    return [serialize_document(doc) for doc in collection.find()]
 
@app.get("/documents/{document_id}")
def get_document_by_id(document_id: str):
    try:
        doc = collection.find_one({"_id": ObjectId(document_id)})
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        return serialize_document(doc)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid document ID format")
 
@app.delete("/delete-documents")
def delete_documents(document_ids: List[str]):
    try:
        for document_id in document_ids:
            doc = collection.find_one({"_id": ObjectId(document_id)})
            if not doc:
                raise HTTPException(status_code=404, detail=f"Document {document_id} not found")
 
            file_path = doc.get("doc_path")
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
 
            collection.delete_one({"_id": ObjectId(document_id)})
 
        return JSONResponse(content={"success": f"{len(document_ids)} documents deleted successfully."}, status_code=status.HTTP_200_OK)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid document ID format")
 
@app.get("/download/testcases/{document_id}")
def download_test_cases_excel(document_id: str):
    try:
        doc = collection.find_one({"_id": ObjectId(document_id)})
        if not doc or "test_case_excel_path" not in doc:
            raise HTTPException(status_code=404, detail="Excel file not found for test cases")
        return FileResponse(doc["test_case_excel_path"], media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=Path(doc["test_case_excel_path"]).name)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid document ID format")
 
@app.get("/download/userstories/{document_id}")
def download_user_stories_excel(document_id: str):
    try:
        doc = collection.find_one({"_id": ObjectId(document_id)})
        if not doc or "user_story_excel_path" not in doc:
            raise HTTPException(status_code=404, detail="Excel file not found for user stories")
        return FileResponse(doc["user_story_excel_path"], media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=Path(doc["user_story_excel_path"]).name)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid document ID format")
 